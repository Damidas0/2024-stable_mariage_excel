from openpyxl import load_workbook
import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askopenfilename


# Fonction pour charger le fichier Excel
def load_excel_file():
    """Ouvre boite de dialogue et charge l'excel séléctionné

    Returns:
        ExcelFile: fichier Excel séléctionné
    """
    Tk().withdraw()  # Masque la fenêtre principale de Tkinter
    file_path = askopenfilename(title="Sélectionnez le fichier Excel")
    return pd.ExcelFile(file_path)


# Fonction pour lire les données des participants
def read_participant_data(excel_file, sheet_name = 'Inscrits'):
    """Lis les inscriptions dans l'Excel

    Args:
        excel_file (ExcelFile): fichier excel à lire

    Returns:
        DataFrame:
    """
    participants = pd.read_excel(excel_file, sheet_name=sheet_name)
    return participants #pd.read_excel(excel_file, sheet_name='Inscrits')


# Fonction pour lire les données des participants
def read_attribution(excel_file, sheet_name = 'Attributions'):
    """Lis les inscriptions dans l'Excel

    Args:
        excel_file (ExcelFile): fichier excel à lire

    Returns:
        DataFrame:
    """
    participants = pd.read_excel(excel_file, sheet_name=sheet_name)#,skiprows=1)
    return participants #pd.read_excel(excel_file, sheet_name='Inscrits')

def read_workshop_data(excel_file):
    """Lis les inscriptions dans l'Excel 

    Args:
        excel_file (ExcelFile): fichier excel à lire

    Returns:
        DataFrame: 
    """
    ateliers = pd.read_excel(excel_file, sheet_name='Ateliers')
    
    return ateliers #pd.read_excel(excel_file, sheet_name='Ateliers')


def delete_attribution(excel_file):
    excel_file.sheet['Attribution'].delete() 
    

# Fonction pour enregistrer les attributions dans une nouvelle feuille Excel
def save_assignments(excel_file, assignments, waitlists):
    
    wb = load_workbook(excel_file)

    # Check if the sheet already exists and remove it
    if 'Attributions' in wb.sheetnames:
        std = wb['Attributions']
        wb.remove(std)
        
    # Check if the sheet already exists and remove it
    if 'Waitlist' in wb.sheetnames:
        std = wb['Waitlist']
        wb.remove(std)
    
    # Save the workbook after removing the sheet
    wb.save(excel_file)

    
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        assignments_data = []
        waitlists_data=[]
        
        
        # Convertir les assignations et listes d'attente en DataFrames pour l'export
        for assigned_workshops, email in assignments.items() : 
            for e in email : 
                #print(assigned_workshops)
                #print(assignments_data)
                if(assigned_workshops != []) :
                    if(e[0] in assignments_data) : 
                        print("djzqijdqjidiqo")
                        assignments_data[e].append(assigned_workshops)
                    assignments_data.append({'Email':e[0], 'Ateliers Attribués': assigned_workshops })
            
        assignments_df = pd.DataFrame(assignments_data) 
        
        
        waitlists_df = pd.DataFrame([(workshop, ', '.join(([e[0] for e in emails]))) for workshop, emails in waitlists.items()if emails!=[]],
                            columns=['Atelier', 'Waitlist'])
        
        
        #for email, assigned_workshops in assignments.items():
        #    assignment_data.append({'Email': email, 'Ateliers Attribués': ', '.join(assigned_workshops)})
        #df = pd.DataFrame(assignment_data)
        
        assignments_df.to_excel(writer, sheet_name='Attributions', index=False)
        waitlists_df.to_excel(writer, sheet_name='Waitlist', index=False)

        #df.to_excel(writer, sheet_name='Attributions', index=False)

